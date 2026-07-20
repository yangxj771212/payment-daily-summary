from __future__ import annotations

import io
import os
import re
from collections import defaultdict
from datetime import date, datetime, time
from html.parser import HTMLParser

import xlrd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

REQUIRED_COLUMNS = ["缴费日期", "业务类别", "预缴金额", "实缴金额"]
OUTPUT_COLUMNS = ["缴费日期按日汇总", "银行存款", "预收报名费", "预收报名费转收入", "收入", "税"]


class InputError(ValueError):
    pass


class ExcelHtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables, self.table, self.row, self.cell = [], None, None, None
        self.colspan = 1

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table" and self.table is None:
            self.table = []
        elif tag == "tr" and self.table is not None:
            self.row = []
        elif tag in ("td", "th") and self.row is not None:
            self.cell = []
            try:
                self.colspan = max(1, int(dict(attrs).get("colspan", "1")))
            except ValueError:
                self.colspan = 1
        elif tag == "br" and self.cell is not None:
            self.cell.append("\n")

    def handle_data(self, data):
        if self.cell is not None:
            self.cell.append(data)

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("td", "th") and self.cell is not None and self.row is not None:
            self.row.append("".join(self.cell).strip())
            self.row.extend([""] * (self.colspan - 1))
            self.cell, self.colspan = None, 1
        elif tag == "tr" and self.row is not None and self.table is not None:
            if self.row:
                self.table.append(self.row)
            self.row = None
        elif tag == "table" and self.table is not None:
            if self.table:
                self.tables.append(self.table)
            self.table = None


def clean_header(value) -> str:
    return re.sub(r"\s+", "", "" if value is None else str(value)).strip()


def html_rows(content: bytes) -> list[list]:
    parser = ExcelHtmlTableParser()
    parser.feed(content.decode("utf-8-sig", errors="replace"))
    if not parser.tables:
        raise InputError("没有在文件中找到数据表。")
    return max(parser.tables, key=lambda rows: len(rows) * max(map(len, rows), default=0))


def xlsx_rows(content: bytes) -> list[list]:
    book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = book.worksheets[0]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def xls_rows(content: bytes) -> list[list]:
    book = xlrd.open_workbook(file_contents=content, on_demand=True)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        values = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                values.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode))
            else:
                values.append(cell.value)
        rows.append(values)
    return rows


def parse_date(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "").replace(".", "-")
    for fmt in ("%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_amount(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").replace(",", "").replace("￥", "").replace("¥", "").strip()
    try:
        return float(text) if text else 0.0
    except ValueError:
        return 0.0


def normalize_rows(rows: list[list]) -> list[dict]:
    rows = [row for row in rows if any(str(v or "").strip() for v in row)]
    if not rows:
        raise InputError("表格没有可读取的数据。")
    header_index = None
    for i, row in enumerate(rows[:30]):
        names = {clean_header(v) for v in row}
        if set(REQUIRED_COLUMNS).issubset(names):
            header_index = i
            break
    if header_index is None:
        raise InputError("缺少必要字段：" + "、".join(REQUIRED_COLUMNS))
    headers = [clean_header(v) for v in rows[header_index]]
    positions = {name: headers.index(name) for name in REQUIRED_COLUMNS}
    output = []
    for row in rows[header_index + 1 :]:
        def value(name):
            idx = positions[name]
            return row[idx] if idx < len(row) else None
        paid_at = parse_date(value("缴费日期"))
        if paid_at is None:
            continue
        output.append({
            "缴费日期": paid_at,
            "业务类别": str(value("业务类别") or "").strip(),
            "预缴金额": parse_amount(value("预缴金额")),
            "实缴金额": parse_amount(value("实缴金额")),
        })
    if not output:
        raise InputError("“缴费日期”列没有有效日期。")
    return output


def read_uploaded_excel(file_storage) -> list[dict]:
    filename = (file_storage.filename or "").lower()
    content = file_storage.read()
    if not content:
        raise InputError("上传的文件为空。")
    try:
        prefix = content[:500].lstrip().lower()
        if b"<html" in prefix or b"<table" in prefix:
            rows = html_rows(content)
        elif filename.endswith((".xlsx", ".xlsm")):
            rows = xlsx_rows(content)
        elif filename.endswith(".xls"):
            rows = xls_rows(content)
        else:
            raise InputError("仅支持 .xls、.xlsx 或 .xlsm 文件。")
        return normalize_rows(rows)
    except InputError:
        raise
    except Exception as exc:
        raise InputError(f"无法读取该 Excel：{exc}") from exc


def calculate_summary(records: list[dict]) -> list[dict]:
    daily = defaultdict(lambda: {"银行存款": 0.0, "预收报名费": 0.0, "预收报名费转收入": 0.0})
    for record in records:
        day = record["缴费日期"].date()
        daily[day]["银行存款"] += record["实缴金额"]
        daily[day]["预收报名费转收入"] += record["预缴金额"]
        if record["业务类别"] == "预缴费":
            daily[day]["预收报名费"] += record["实缴金额"]
    result = []
    for day in sorted(daily):
        values = daily[day]
        income = (values["银行存款"] - values["预收报名费"] + values["预收报名费转收入"]) / 1.03
        result.append({"缴费日期按日汇总": day, **values, "收入": income, "税": income * 0.03})
    return result


def make_workbook(summary: list[dict], records: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "缴费日期按日汇总"
    detail = wb.create_sheet("缴费明细")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(OUTPUT_COLUMNS)
    for cell in ws[1]:
        cell.fill, cell.border, cell.alignment = header_fill, border, Alignment(horizontal="left")
    for item in summary:
        ws.append([item[c] for c in OUTPUT_COLUMNS])
    total_row = ws.max_row + 1
    ws.cell(total_row, 1, "合计")
    for col in range(2, 7):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f"=SUM({letter}2:{letter}{total_row - 1})")
    for cell in ws[total_row]:
        cell.fill, cell.font, cell.border = total_fill, Font(bold=True, color="375623"), border
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = "yyyy-mm-dd"
        for cell in row[1:]:
            cell.number_format, cell.border = "#,##0.00", border
    for idx, width in enumerate([22, 17, 18, 23, 17, 15], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes, ws.auto_filter.ref = "A2", f"A1:F{ws.max_row - 1}"

    detail_headers = ["缴费日期", "业务类别", "预缴金额", "实缴金额"]
    detail.append(detail_headers)
    for cell in detail[1]:
        cell.fill, cell.font, cell.border = header_fill, Font(bold=True), border
    for item in records:
        detail.append([item[c] for c in detail_headers])
    for row in detail.iter_rows(min_row=2):
        row[0].number_format = "yyyy-mm-dd hh:mm"
        for cell in row:
            cell.border = border
        for cell in row[2:]:
            cell.number_format = "#,##0.00"
    for i, width in enumerate([20, 16, 18, 18], 1):
        detail.column_dimensions[get_column_letter(i)].width = width
    detail.freeze_panes, detail.auto_filter.ref = "A2", detail.dimensions
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/api/process")
def process_file():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "请选择 Excel 文件。"}), 400
    try:
        records = read_uploaded_excel(request.files["file"])
        output = make_workbook(calculate_summary(records), records)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(output, as_attachment=True, download_name=f"缴费日期会计汇总_{stamp}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except InputError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception as exc:
        app.logger.exception("Processing failed")
        return jsonify({"ok": False, "error": f"处理失败：{exc}"}), 500


@app.get("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
